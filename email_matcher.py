import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('sample_employee_data.xlsx')
sheet1 = wb['Sheet1']
template = wb['Employee Template']

# Step 1: Build a lookup dictionary from Sheet1 (Employee ID -> Email)
email_lookup = {}
for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, values_only=True):
    emp_id = str(row[0]).strip() if row[0] else ''
    email = row[1]
    if emp_id and email:
        email_lookup[emp_id] = email

print(f'Email lookup entries: {len(email_lookup)}')

# Step 2: Match Employee IDs and fill Column H in Employee Template
matched = 0
for row_num in range(2, template.max_row + 1):
    emp_id_cell = template.cell(row=row_num, column=2)  # Column B = Employee ID
    emp_id = str(emp_id_cell.value).strip() if emp_id_cell.value else ''
    if emp_id in email_lookup:
        template.cell(row=row_num, column=8).value = email_lookup[emp_id]  # Column H = Email
        matched += 1

# Step 3: Check for unmatched IDs from Sheet1
template_ids = set()
for row in template.iter_rows(min_row=2, max_row=template.max_row, values_only=True):
    template_ids.add(str(row[1]).strip() if row[1] else '')

unmatched = [eid for eid in email_lookup if eid not in template_ids]

print(f'Matched and filled: {matched}')
print(f'Sheet1 IDs not found in template: {len(unmatched)}')
if unmatched:
    print('Unmatched IDs:', unmatched)

# Step 4: Save the file
wb.save('sample_employee_data.xlsx')
print('File saved successfully!')
